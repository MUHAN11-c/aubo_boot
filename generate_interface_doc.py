#!/usr/bin/env python3
"""自动发现 ROS2 工作空间接口并导出 接口明细.xlsx

扫描 src/ 下的 .msg/.srv/.action 及 C++/Python 源码，
自动提取话题/服务/动作/参数，参照参考表格样式分4页输出。
用法: python3 generate_interface_doc.py [workspace_src_dir]
"""

import os, re, sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "/home/mu/IVG2.0/aubo_ros2_ws/src"
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "接口明细.xlsx")
SKIP_DIRS = {"install", "build", "log", "examples", "robotwebtools", "3rdparty", "third_party", ".git"}

# ── 样式 ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_FONT = Font(name="微软雅黑", size=10, color="FF000000")
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True, color="FF000000")
PARENT_A_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
PARENT_B_I_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CHILD_FILL_A = PatternFill()
CHILD_FILL_B = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="FF000000"), right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"), bottom=Side(style="thin", color="FF000000"),
)
COLUMNS = ["功能包", "名称/说明", "类型", "变量名", "变量类型", "所属包/模块", "C++类型", "Python类型", "说明", "回首页"]
COL_WIDTHS = [25.0, 21.4, 24.3, 17.6, 20.0, 26.3, 41.5, 46.1, 32.8, 11.3]
SHEETS = ["参数 (Parameters)", "话题 (Topics)", "服务 (Services)", "动作 (Actions)"]

# ── 工具 ──────────────────────────────────────────────────────────────────
def skip(path):
    return bool(set(os.path.normpath(path).split(os.sep)) & SKIP_DIRS)

# ── 包名缓存：目录路径 → package.xml <name> ──
_pkg_cache = {}

def _build_pkg_cache(src_dir):
    """扫描所有 package.xml，建立 目录路径 → 包名 的映射"""
    if _pkg_cache:
        return
    for root, dirs, files in os.walk(src_dir):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        if "package.xml" in files:
            try:
                with open(os.path.join(root, "package.xml"), "r", encoding="utf-8") as f:
                    content = f.read()
                m = re.search(r"<name>([^<]+)</name>", content)
                if m:
                    _pkg_cache[os.path.normpath(root)] = m.group(1)
            except:
                pass

def pkg_name(filepath):
    """根据文件路径查找所属 ROS2 功能包名称（向上查找 package.xml）"""
    path = os.path.normpath(os.path.dirname(os.path.abspath(filepath)))
    # 向上遍历目录树，直到找到 package.xml 或根目录
    while path and path != os.sep:
        if path in _pkg_cache:
            return _pkg_cache[path]
        parent = os.path.dirname(path)
        if parent == path:
            break
        path = parent
    # fallback: 旧逻辑
    parts = os.path.normpath(filepath).split(os.sep)
    for i, p in enumerate(parts):
        if p in ("msg", "srv", "action") and i > 0:
            return parts[i-1]
    return os.path.basename(os.path.dirname(os.path.dirname(filepath)))

def first_comment_line(filepath):
    """读取文件的第一个 # 注释行"""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if s.startswith("#"):
                    return s.lstrip("#").strip()
                if s and not s.startswith("#"):
                    break
    except: pass
    return ""

def inline_comment(line):
    if "#" in line:
        return line.split("#", 1)[1].strip()
    return ""

# ── 类型转换 ──────────────────────────────────────────────────────────────
PRIM_CPP = {"bool":"bool","int8":"int8_t","uint8":"uint8_t","int16":"int16_t","uint16":"uint16_t",
            "int32":"int32_t","uint32":"uint32_t","int64":"int64_t","uint64":"uint64_t",
            "float32":"float","float64":"double","string":"std::string","char":"unsigned char","byte":"uint8_t"}
PRIM_PY = {"bool":"bool","int8":"int","uint8":"int","int16":"int","uint16":"int",
           "int32":"int","uint32":"int","int64":"int","uint64":"int",
           "float32":"float","float64":"float","string":"str","char":"int","byte":"int"}
KNOWN_MSGS = {
    "Header":"std_msgs::msg::Header","Pose":"geometry_msgs::msg::Pose",
    "Point":"geometry_msgs::msg::Point","Quaternion":"geometry_msgs::msg::Quaternion",
    "Vector3":"geometry_msgs::msg::Vector3","Image":"sensor_msgs::msg::Image",
    "JointState":"sensor_msgs::msg::JointState","JointTrajectory":"trajectory_msgs::msg::JointTrajectory",
    "JointTrajectoryPoint":"trajectory_msgs::msg::JointTrajectoryPoint",
    "PoseArray":"geometry_msgs::msg::PoseArray","PointCloud2":"sensor_msgs::msg::PointCloud2",
    "CameraInfo":"sensor_msgs::msg::CameraInfo","TFMessage":"tf2_msgs::msg::TFMessage",
    "Marker":"visualization_msgs::msg::Marker","TransformStamped":"geometry_msgs::msg::TransformStamped",
    "PointField":"sensor_msgs::msg::PointField","Float32MultiArray":"std_msgs::msg::Float32MultiArray",
    "Int32MultiArray":"std_msgs::msg::Int32MultiArray","String":"std_msgs::msg::String",
    "UInt8":"std_msgs::msg::UInt8","Int32":"std_msgs::msg::Int32","Bool":"std_msgs::msg::Bool",
}

def ros_to_cpp(ftype, pkg=""):
    base = ftype.rstrip("[]"); is_arr = base != ftype
    if base in PRIM_CPP:
        c = PRIM_CPP[base]; return f"std::vector<{c}>" if is_arr else c
    if base in KNOWN_MSGS:
        c = KNOWN_MSGS[base]; return f"std::vector<{c}>" if is_arr else c
    if "/" in ftype:
        parts = ftype.split("/"); c = f"{parts[0]}::msg::{parts[1]}"
        return f"std::vector<{c}>" if is_arr else c
    c = f"{pkg}::msg::{base}"
    return f"std::vector<{c}>" if is_arr else c

def ros_to_py(ftype, pkg=""):
    base = ftype.rstrip("[]"); is_arr = base != ftype
    if base in PRIM_PY:
        p = PRIM_PY[base]; return f"list[{p}]" if is_arr else p
    for short, cpp in KNOWN_MSGS.items():
        if base == short:
            pfx = cpp.split("::")[0]; p = f"{pfx}.msg.{short}"
            return f"list[{p}]" if is_arr else p
    if "/" in ftype:
        parts = ftype.split("/"); p = f"{parts[0]}.msg.{parts[1]}"
        return f"list[{p}]" if is_arr else p
    p = f"{pkg}.msg.{base}"; return f"list[{p}]" if is_arr else p

def cpp_type_to_ros(cpp):
    """demo_interface::msg::RobotStatus → demo_interface/msg/RobotStatus"""
    t = cpp.strip()
    for pfx in ["std_msgs::msg::","geometry_msgs::msg::","sensor_msgs::msg::",
                "trajectory_msgs::msg::","control_msgs::action::","tf2_msgs::msg::",
                "visualization_msgs::msg::","nav_msgs::msg::"]:
        if t.startswith(pfx):
            pkg, _, name = pfx.rstrip(":").split("::")
            return f"{pkg}/msg/{t[len(pfx):]}" if "msg" in pfx else f"{pkg}/action/{t[len(pfx):]}"
    if "::msg::" in t:
        return t.replace("::msg::", "/msg/")
    if "::srv::" in t:
        return t.replace("::srv::", "/srv/")
    if "::action::" in t:
        return t.replace("::action::", "/action/")
    return t

# ── 消息/服务/动作 文件解析 ────────────────────────────────────────────────

def parse_msg(filepath):
    """返回 (描述, [(变量名, 变量类型, 包, C++类型, Python类型, 说明), ...])"""
    pkg = pkg_name(filepath)
    desc = first_comment_line(filepath) or os.path.basename(filepath)[:-4]
    fields = []
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"): continue
            if "=" in s: continue
            parts = s.split()
            if len(parts) >= 2:
                ftype, fname = parts[0], parts[1]
                fields.append((fname, ftype, pkg, ros_to_cpp(ftype, pkg), ros_to_py(ftype, pkg), inline_comment(s)))
    return desc, fields

def parse_srv(filepath):
    """返回 (描述, [(变量名, 变量类型, 包, C++类型, Python类型, 说明), ...], response_fields)"""
    pkg = pkg_name(filepath)
    desc = first_comment_line(filepath) or os.path.basename(filepath)[:-4]
    req, resp = [], []
    target = req
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s: continue
            if s.startswith("#"): continue
            if s == "---": target = resp; continue
            parts = s.split()
            if len(parts) >= 2:
                ftype, fname = parts[0], parts[1]
                target.append((fname, ftype, pkg, ros_to_cpp(ftype, pkg), ros_to_py(ftype, pkg), inline_comment(s)))
    return desc, req, resp

def parse_action(filepath):
    """返回 (描述, goal_fields, result_fields, feedback_fields)"""
    pkg = pkg_name(filepath)
    desc = first_comment_line(filepath) or os.path.basename(filepath)[:-7]
    sections = [[], [], []]; idx = 0
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s: continue
            if s.startswith("#"): continue
            if s == "---": idx += 1; continue
            parts = s.split()
            if len(parts) >= 2:
                ftype, fname = parts[0], parts[1]
                sections[idx].append((fname, ftype, pkg, ros_to_cpp(ftype, pkg), ros_to_py(ftype, pkg), inline_comment(s)))
    return desc, sections[0], sections[1], sections[2]

# ── 源码扫描 ──────────────────────────────────────────────────────────────

def scan_cpp(src_dir):
    pubs, subs, svcs, params = [], [], [], []
    for root, dirs, files in os.walk(src_dir):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        for fn in files:
            if not fn.endswith((".cpp",".hpp",".h",".cc",".cxx")): continue
            fp = os.path.join(root, fn); pkg = pkg_name(fp)
            try:
                with open(fp,"r",encoding="utf-8",errors="ignore") as f: content = f.read()
            except: continue

            # ── 先构建 param_key → default_value 映射（用于解析变量话题/服务名）──
            param_defaults = {}  # "status_topic" → "grasp_place_status"
            for m in re.finditer(r'declare_parameter\s*(?:<[^>]+>)?\s*\(\s*"([^"]+)"\s*,\s*std::string\s*\(\s*"([^"]*)"\s*\)', content):
                param_defaults[m.group(1)] = m.group(2)
            # var_ = get_parameter("key").as_string() → var_ → default
            var_to_default = {}  # "status_topic_" → "grasp_place_status"
            for m in re.finditer(r'(\w+)\s*=\s*get_parameter\s*\(\s*"([^"]+)"\s*\)\s*\.\s*as_string\s*\(\s*\)', content):
                if m.group(2) in param_defaults:
                    var_to_default[m.group(1)] = param_defaults[m.group(2)]

            # ── 字面量话题/服务名 ──
            for m in re.finditer(r'create_publisher\s*<\s*([^>]+)\s*>\s*\(\s*"([^"]+)"\s*,\s*([^,)]+(?:\s*\([^)]*\)[^,)]*)?)', content):
                pubs.append((m.group(2), m.group(1).strip(), m.group(3), pkg))
            for m in re.finditer(r'create_subscription\s*<\s*([^>]+)\s*>\s*\(\s*"([^"]+)"\s*,\s*([^,)]+(?:\s*\([^)]*\)[^,)]*)?)', content):
                subs.append((m.group(2), m.group(1).strip(), m.group(3), pkg))
            for m in re.finditer(r'create_service\s*<\s*([^>]+)\s*>\s*\(\s*"([^"]+)"', content):
                svcs.append((m.group(2), m.group(1).strip(), pkg))

            # ── 变量话题/服务名（从 declare_parameter 解析默认值）──
            for m in re.finditer(r'create_publisher\s*<\s*([^>]+)\s*>\s*\(\s*(\w+)\s*,\s*([^,)]+(?:\s*\([^)]*\)[^,)]*)?)', content):
                if m.group(2) in var_to_default:
                    pubs.append((var_to_default[m.group(2)], m.group(1).strip(), m.group(3), pkg))
            for m in re.finditer(r'create_subscription\s*<\s*([^>]+)\s*>\s*\(\s*(\w+)\s*,\s*([^,)]+(?:\s*\([^)]*\)[^,)]*)?)', content):
                if m.group(2) in var_to_default:
                    subs.append((var_to_default[m.group(2)], m.group(1).strip(), m.group(3), pkg))
            for m in re.finditer(r'create_service\s*<\s*([^>]+)\s*>\s*\(\s*(\w+)\s*,', content):
                if m.group(2) in var_to_default:
                    svcs.append((var_to_default[m.group(2)], m.group(1).strip(), pkg))

            for line in content.split("\n"):
                s = line.strip()
                if s.startswith("//"): continue
                for m in re.finditer(r'declare_parameter\s*(?:<([^>]+)>)?\s*\(\s*"([^"]+)"\s*,\s*((?:[^()]|\((?:[^()]*)\))*?)\)', s):
                    ptype = (m.group(1) or "").strip()
                    pname = m.group(2)
                    pdef = m.group(3).strip().rstrip(",").strip()
                    pdef = re.sub(r'std::string\s*\(\s*"([^"]*)"\s*\)', r'"\1"', pdef)
                    pdef = re.sub(r'std::string\s*\(\s*"', '', pdef)
                    pdef = pdef.replace('{','').replace('}','').strip('" ')
                    pdef = pdef.rstrip('f')
                    if not ptype:
                        if pdef in ("true","false"): ptype = "bool"
                        elif re.match(r'^-?\d+\.\d+$', pdef): ptype = "double"
                        elif re.match(r'^-?\d+$', pdef): ptype = "int"
                        elif pdef.startswith('"'): ptype = "string"
                        else: ptype = "string"
                    params.append((pname, pdef, ptype, pkg))
    return pubs, subs, svcs, params

def scan_py(src_dir):
    pubs, subs, svcs, params = [], [], [], []
    for root, dirs, files in os.walk(src_dir):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        for fn in files:
            if not fn.endswith(".py"): continue
            fp = os.path.join(root, fn); pkg = pkg_name(fp)
            try:
                with open(fp,"r",encoding="utf-8",errors="ignore") as f: txt = f.read()
            except: continue

            for m in re.finditer(r'create_publisher\s*\(\s*([^,]+)\s*,\s*["\']([^"\']+)["\']', txt):
                t = m.group(1).strip()
                if t.startswith("self."): t = t[5:]
                pubs.append((m.group(2), t, "10", pkg))
            for m in re.finditer(r'create_subscription\s*\(\s*([^,]+)\s*,\s*["\']([^"\']+)["\']', txt):
                t = m.group(1).strip()
                if t.startswith("self."): t = t[5:]
                subs.append((m.group(2), t, "10", pkg))
            for m in re.finditer(r'create_service\s*\(\s*([^,]+)\s*,\s*["\']([^"\']+)["\']', txt):
                t = m.group(1).strip()
                if t.startswith("self."): t = t[5:]
                svcs.append((m.group(2), t, pkg))
            for m in re.finditer(r'declare_parameter\s*\(\s*["\']([^"\']+)["\']\s*,\s*([^)]+)\)', txt):
                params.append((m.group(1), m.group(2).strip().rstrip(",").strip(), "", pkg))
    return pubs, subs, svcs, params

# ── 核心：匹配运行时接口与类型定义 ────────────────────────────────────────

def build_type_map(src_dir):
    """构建 ROS类型路径 → (描述, 字段列表) 的映射"""
    type_map = {}  # "pkg/msg/Name" → (desc, fields)
    for root, dirs, files in os.walk(src_dir):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        for fn in files:
            fp = os.path.join(root, fn)
            if fn.endswith(".msg"):
                pkg = pkg_name(fp)
                desc, fields = parse_msg(fp)
                type_map[f"{pkg}/msg/{fn[:-4]}"] = (desc, fields)
            elif fn.endswith(".srv"):
                pkg = pkg_name(fp)
                desc, req, resp = parse_srv(fp)
                type_map[f"{pkg}/srv/{fn[:-4]}"] = (desc, req, resp)
            elif fn.endswith(".action"):
                pkg = pkg_name(fp)
                desc, goal, result, feedback = parse_action(fp)
                type_map[f"{pkg}/action/{fn[:-7]}"] = (desc, goal, result, feedback)
    return type_map

def resolve_type(cpp_type, type_map):
    """将 C++ 类型字符串解析为 type_map 中的 key"""
    t = cpp_type.strip()
    # demo_interface::msg::RobotStatus → demo_interface/msg/RobotStatus
    if "::msg::" in t:
        parts = t.split("::msg::"); return f"{parts[0]}/msg/{parts[1]}"
    if "::srv::" in t:
        parts = t.split("::srv::"); return f"{parts[0]}/srv/{parts[1]}"
    if "::action::" in t:
        parts = t.split("::action::"); return f"{parts[0]}/action/{parts[1]}"
    # Python: RobotStatus → demo_interface/msg/RobotStatus
    for key in type_map:
        if key.endswith(f"/{t}") or key.endswith(f"/msg/{t}") or key.endswith(f"/srv/{t}"):
            return key
    return None

# ── Excel 写入 ────────────────────────────────────────────────────────────

def wb_header(ws):
    for ci, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def wb_parent(ws, row, pkg, name, desc, typ, vname, vtype, cpp, py, note):
    # A=功能包 B=名称+说明
    bval = f"{name} — {desc}" if desc else name
    vals = [pkg, bval, typ, vname, vtype, pkg, cpp, py, note, None]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font, c.alignment, c.border = DATA_FONT, DATA_ALIGN, THIN_BORDER
        c.fill = PARENT_A_FILL if ci == 1 else PARENT_B_I_FILL

def wb_child(ws, row, vname, vtype, pkg, cpp, py, note, alt):
    # B列=中文说明(取注释,无注释则用字段名); C列=字段名; D列=字段类型
    desc = note if note else vname
    vals = [None, desc, vtype, vname, pkg, cpp, py, note, None, None]
    fill = CHILD_FILL_B if alt else CHILD_FILL_A
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font, c.alignment, c.border = DATA_FONT, DATA_ALIGN, THIN_BORDER
        c.fill = fill

def wb_sep(ws, row, label, alt):
    vals = [None, None, label, None, None, None, None, None, None, None]
    fill = CHILD_FILL_B if alt else CHILD_FILL_A
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font, c.alignment, c.border = BOLD_FONT, DATA_ALIGN, THIN_BORDER
        c.fill = fill

# ── 主流程 ────────────────────────────────────────────────────────────────

def main():
    print(f"扫描: {SRC}")
    _build_pkg_cache(SRC)  # 预构建 package.xml 目录→包名 映射
    type_map = build_type_map(SRC)
    cpp_pubs, cpp_subs, cpp_svcs, cpp_params = scan_cpp(SRC)
    py_pubs, py_subs, py_svcs, py_params = scan_py(SRC)

    # 去重
    def dedup(lst):
        seen = set(); out = []
        for item in lst:
            key = (item[0], item[1])  # name + type
            if key not in seen: seen.add(key); out.append(item)
        return out

    cpp_pubs, cpp_subs = dedup(cpp_pubs), dedup(cpp_subs)
    cpp_svcs = dedup(cpp_svcs); cpp_params = dedup(cpp_params)
    py_pubs, py_subs = dedup(py_pubs), dedup(py_subs)
    py_svcs = dedup(py_svcs); py_params = dedup(py_params)

    print(f"  消息类型: {sum(1 for k in type_map if '/msg/' in k)} 个")
    print(f"  服务类型: {sum(1 for k in type_map if '/srv/' in k)} 个")
    print(f"  动作类型: {sum(1 for k in type_map if '/action/' in k)} 个")
    print(f"  C++: pub={len(cpp_pubs)} sub={len(cpp_subs)} svc={len(cpp_svcs)} param={len(cpp_params)}")
    print(f"  Python: pub={len(py_pubs)} sub={len(py_subs)} svc={len(py_svcs)} param={len(py_params)}")

    # ── 组装数据（含包名，按包排序）──
    # 话题: [(包名, 话题名, 中文说明, 字段列表)]
    topics = []
    seen_topics = set()
    for topic, cpp_type, qos, scan_pkg in cpp_pubs + cpp_subs + py_pubs + py_subs:
        if topic in seen_topics:
            continue
        type_key = resolve_type(cpp_type, type_map)
        if type_key and type_key in type_map:
            info = type_map[type_key]
            if isinstance(info[1], list) and info[1] and isinstance(info[1][0], tuple):
                desc, fields = info
                # 包名来自类型定义的包（如 demo_interface/msg/RobotStatus → demo_interface）
                pkg = type_key.split("/")[0] if "/" in type_key else scan_pkg
                topics.append((pkg, topic, desc, fields))
                seen_topics.add(topic)
        else:
            # 标准ROS类型（如 std_msgs::msg::String）— 无字段展开，仍记录话题
            desc = cpp_type.split("::")[-1] if "::" in cpp_type else cpp_type
            topics.append((scan_pkg, topic, cpp_type, []))
            seen_topics.add(topic)
    topics.sort(key=lambda x: (x[0], x[1]))

    # 服务: [(包名, 服务名, 中文说明, req, resp)]
    services = []
    seen_svcs = set()
    for svc, cpp_type, scan_pkg in cpp_svcs + py_svcs:
        if svc in seen_svcs:
            continue
        type_key = resolve_type(cpp_type, type_map)
        if type_key and type_key in type_map:
            info = type_map[type_key]
            if isinstance(info, tuple) and len(info) == 3:
                desc, req, resp = info
                pkg = type_key.split("/")[0] if "/" in type_key else scan_pkg
                services.append((pkg, svc, desc, req, resp))
                seen_svcs.add(svc)
        else:
            # 标准ROS类型或无法解析 — 仍记录服务
            desc = cpp_type.split("::")[-1] if "::" in cpp_type else cpp_type
            services.append((scan_pkg, svc, cpp_type, [], []))
            seen_svcs.add(svc)
    services.sort(key=lambda x: (x[0], x[1]))

    # 参数: [(包名, 参数名, 默认值, 类型)]
    param_map = {}
    for pname, pdef, ptype, pkg in cpp_params + py_params:
        if pname not in param_map:
            param_map[pname] = (pkg, pname, pdef, ptype)
    params_sorted = sorted(param_map.values(), key=lambda x: (x[0], x[1]))

    # 动作: [(包名, 动作名, 中文说明, goal, result, feedback)]
    actions = []
    for key, info in type_map.items():
        if '/action/' in key:
            if isinstance(info, tuple) and len(info) == 4:
                desc, goal, result, feedback = info
                pkg = key.split("/")[0]
                actions.append((pkg, key, desc, goal, result, feedback))
    actions.sort(key=lambda x: (x[0], x[1]))

    # ── 写入 Excel ──
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_objs = {}
    row_ptrs = {}
    for title in SHEETS:
        ws = wb.create_sheet(title=title)
        wb_header(ws)
        sheet_objs[title] = ws
        row_ptrs[title] = 2

    # -- 参数 --
    ws = sheet_objs["参数 (Parameters)"]; row = row_ptrs["参数 (Parameters)"]
    for pkg, pname, pdef, ptype in params_sorted:
        cpp_t = {"bool":"bool","int":"int","double":"double","string":"std::string","float":"float"}.get(ptype, "std::string")
        py_t = {"bool":"bool","int":"int","double":"float","string":"str","float":"float"}.get(ptype, "str")
        wb_parent(ws, row, pkg, pname, "", "Parameter", "—", ptype, cpp_t, py_t, f"默认值: {pdef}")
        row += 1
    row_ptrs["参数 (Parameters)"] = row

    # -- 话题 --
    ws = sheet_objs["话题 (Topics)"]; row = row_ptrs["话题 (Topics)"]
    for pkg, topic, desc, fields in topics:
        if not fields: continue
        first = fields[0]
        wb_parent(ws, row, pkg, topic, desc, "Topic", first[0], first[1], first[3], first[4], first[5])
        row += 1; alt = False
        for fi in range(1, len(fields)):
            fname, ftype, fpkg, cpp, py, note = fields[fi]
            wb_child(ws, row, fname, ftype, fpkg, cpp, py, note, alt)
            row += 1; alt = not alt
    row_ptrs["话题 (Topics)"] = row

    # -- 服务 --
    ws = sheet_objs["服务 (Services)"]; row = row_ptrs["服务 (Services)"]
    for pkg, svc, desc, req, resp in services:
        wb_parent(ws, row, pkg, svc, desc, "Service", "Request", None, None, None, None)
        row += 1; alt = False
        for fname, ftype, fpkg, cpp, py, note in req:
            wb_child(ws, row, fname, ftype, fpkg, cpp, py, note, alt)
            row += 1; alt = not alt
        wb_sep(ws, row, "Response", alt); row += 1; alt = not alt
        for fname, ftype, fpkg, cpp, py, note in resp:
            wb_child(ws, row, fname, ftype, fpkg, cpp, py, note, alt)
            row += 1; alt = not alt
    row_ptrs["服务 (Services)"] = row

    # -- 动作 --
    ws = sheet_objs["动作 (Actions)"]; row = row_ptrs["动作 (Actions)"]
    for pkg, act_name, desc, goal, result, feedback in actions:
        wb_parent(ws, row, pkg, act_name, desc, "Action", "Goal", None, None, None, None)
        row += 1; alt = False
        for fname, ftype, fpkg, cpp, py, note in goal:
            wb_child(ws, row, fname, ftype, fpkg, cpp, py, note, alt)
            row += 1; alt = not alt
        for label, section in [("Result", result), ("Feedback", feedback)]:
            wb_sep(ws, row, label, alt); row += 1; alt = not alt
            for fname, ftype, fpkg, cpp, py, note in section:
                wb_child(ws, row, fname, ftype, fpkg, cpp, py, note, alt)
                row += 1; alt = not alt
    row_ptrs["动作 (Actions)"] = row

    for ws in sheet_objs.values():
        ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"\n已生成: {OUTPUT}")
    for title in SHEETS:
        print(f"  {title}: {sheet_objs[title].max_row - 1} 行")

if __name__ == "__main__":
    main()
